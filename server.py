"""
WorkPDF — FastAPI server
Serves the SPA, handles PDF upload, and exposes uploaded files for preview.
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from pathlib import Path
import shutil
import uuid
import csv
import io
import zipfile
import tempfile

from PDFWrite.pdfwrite import makeconst

# ── Paths ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# ── App ────────────────────────────────────────────────
app = FastAPI(title="WorkPDF")

# Static files (css, js, fonts)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

# Serve uploaded PDFs so the frontend can fetch them for preview
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")


# ── Routes ─────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    """Serve the main page."""
    html_path = BASE_DIR / "templates" / "index.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@app.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Receive a PDF file, save it with a unique name, and return the URL
    so the frontend can render a preview with pdf.js.
    """
    # Validate content type
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF.")

    # Generate a unique filename to avoid collisions
    ext = Path(file.filename).suffix or ".pdf"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    dest = UPLOAD_DIR / unique_name

    # Save file
    with open(dest, "wb") as buf:
        shutil.copyfileobj(file.file, buf)

    return {
        "ok": True,
        "filename": unique_name,
        "url": f"/uploads/{unique_name}",
    }


@app.post("/upload-list")
async def upload_list(file: UploadFile = File(...)):
    """
    Receive a CSV or XLSX list file, extract all elements from the
    first column (skipping the header row) as a list of strings —
    the exact format that makeconst(nombres, ...) expects.
    Also returns the first element separately for the preview.
    """
    filename = file.filename.lower()
    content = await file.read()

    nombres: list[str] = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")  # handle BOM
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            # Skip header row (index 0), take all data rows
            for row in rows[1:]:
                if row and row[0].strip():
                    nombres.append(row[0].strip())

        elif filename.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, max_col=1, values_only=True))
            wb.close()
            for row in rows:
                val = row[0]
                if val is not None and str(val).strip():
                    nombres.append(str(val).strip())
        else:
            raise HTTPException(
                status_code=400,
                detail="Solo se aceptan archivos CSV o XLSX."
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {exc}")

    if not nombres:
        raise HTTPException(status_code=400, detail="No se encontraron datos en el archivo.")

    return {
        "ok": True,
        "nombres": nombres,
        "first_element": nombres[0],
        "count": len(nombres),
    }


# ── Generate constancias ───────────────────────────────

class GenerateRequest(BaseModel):
    nombres: list[str]
    cord_x: float
    cord_y: float
    formato: str
    font: str
    size: float
    pdf_filename: str   # filename in /uploads/


@app.post("/generate")
async def generate(req: GenerateRequest):
    """
    Call makeconst with the collected inputs, zip the generated
    PDFs, and return the zip as a downloadable file.
    """
    # Validate the uploaded PDF exists
    plantilla_path = UPLOAD_DIR / req.pdf_filename
    if not plantilla_path.is_file():
        raise HTTPException(status_code=400, detail="PDF plantilla no encontrado. Sube el PDF primero.")

    if not req.nombres:
        raise HTTPException(status_code=400, detail="No se proporcionó una lista de nombres.")

    # Create a unique temp output directory
    job_id = uuid.uuid4().hex[:12]
    output_dir = BASE_DIR / "Salidas" / job_id

    try:
        # Run makeconst
        archivos = makeconst(
            nombres=req.nombres,
            cord_x=req.cord_x,
            cord_y=req.cord_y,
            input_pagesize=req.formato,
            font=req.font,
            size=req.size,
            plantilla_path=str(plantilla_path),
            output_dir=str(output_dir),
        )

        if not archivos:
            raise HTTPException(status_code=500, detail="No se generaron archivos.")

        # Zip all generated PDFs
        zip_path = BASE_DIR / "Salidas" / f"Constancias_{job_id}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for filepath in archivos:
                zf.write(filepath, arcname=Path(filepath).name)

        # Clean up the individual PDFs folder
        shutil.rmtree(output_dir, ignore_errors=True)

        # Return zip for download
        return FileResponse(
            path=str(zip_path),
            media_type="application/zip",
            filename="Constancias.zip",
            # Delete the zip after it's been sent
            background=None,
        )

    except HTTPException:
        raise
    except Exception as exc:
        # Clean up on error
        shutil.rmtree(output_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Error al generar constancias: {exc}")


# ── Run ────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="127.0.0.1", port=3900, reload=True)
